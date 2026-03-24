import hashlib
import os
import sys
import re
import uuid
import time
import socket
import tempfile
import numpy as np
from datetime import datetime, timedelta
from functools import wraps

import cv2
import jwt
import chromadb
from deepface import DeepFace
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
import mysql.connector
from mysql.connector import Error
import pandas as pd
import qrcode
from io import BytesIO
import base64

from student_list import ALL_STUDENTS

# Fix Windows console encoding for DeepFace emoji logs
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ── Config ────────────────────────────────────────────────────────────────────
DATA_DIR    = os.path.dirname(__file__)
CHROMA_DIR  = os.path.join(DATA_DIR, "chroma_db")
MODEL_NAME  = "Facenet512"
DETECTOR    = "retinaface"
SIMILARITY_THRESHOLD = 0.50
QR_TTL_SEC  = 120

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

_local_ip    = get_local_ip()
FRONTEND_URL = os.environ.get("FRONTEND_URL", f"http://{_local_ip}:5173")
BACKEND_URL  = os.environ.get("BACKEND_URL",  f"http://{_local_ip}:5000")
print(f"[CONFIG] Local IP   : {_local_ip}")
print(f"[CONFIG] Frontend   : {FRONTEND_URL}")
print(f"[CONFIG] Backend    : {BACKEND_URL}")

app = Flask(__name__)
app.config["SECRET_KEY"] = "attendance_secret_key_2026"
CORS(app)

DB_CONFIG = {
    "host":     "localhost",
    "user":     "root",
    "password": "root",
    "database": "attendance_system",
}

QR_STORE: dict = {}

# ── ChromaDB ──────────────────────────────────────────────────────────────────
chroma_client = chromadb.PersistentClient(path=CHROMA_DIR)


def get_collection():
    try:
        return chroma_client.get_collection("face_embeddings")
    except Exception:
        return None


# ── Helpers ───────────────────────────────────────────────────────────────────
def l2_normalize(vec):
    arr = np.array(vec, dtype=np.float32)
    norm = np.linalg.norm(arr)
    return (arr / norm).tolist() if norm > 0 else arr.tolist()


def get_db_connection():
    try:
        return mysql.connector.connect(**DB_CONFIG)
    except Error as e:
        print(f"[DB ERROR] {e}")
        return None


def init_db():
    conn = get_db_connection()
    if not conn:
        return
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS faculty (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            email VARCHAR(255) UNIQUE NOT NULL,
            password VARCHAR(255) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INT AUTO_INCREMENT PRIMARY KEY,
            student_id VARCHAR(50) NOT NULL,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            faculty_id INT,
            FOREIGN KEY (faculty_id) REFERENCES faculty(id)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS qr_attendance (
            id INT AUTO_INCREMENT PRIMARY KEY,
            student_id VARCHAR(50) NOT NULL,
            qr_token VARCHAR(100) NOT NULL,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            faculty_id INT,
            FOREIGN KEY (faculty_id) REFERENCES faculty(id)
        )
    """)
    try:
        cur.execute("ALTER TABLE attendance ADD COLUMN method VARCHAR(20) DEFAULT 'face'")
        conn.commit()
        print("[DB] Added 'method' column to attendance table")
    except Exception:
        pass
    try:
        cur.execute("ALTER TABLE attendance ADD COLUMN session_id VARCHAR(36) DEFAULT NULL")
        conn.commit()
        print("[DB] Added 'session_id' column to attendance table")
    except Exception:
        pass
    conn.commit()
    cur.close()
    conn.close()


def validate_password(password):
    if len(password) < 6:
        return False, "Password must be at least 6 characters"
    if not re.search(r'[!@#$%^&*(),.?":{}|<>]', password):
        return False, "Password must contain at least one special character"
    return True, ""


def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        if not auth:
            return jsonify({"message": "Token is missing"}), 401
        try:
            tok = auth[7:] if auth.startswith("Bearer ") else auth
            data = jwt.decode(tok, app.config["SECRET_KEY"], algorithms=["HS256"])
            current_user_id = data["user_id"]
        except Exception:
            return jsonify({"message": "Token is invalid"}), 401
        return f(current_user_id, *args, **kwargs)
    return decorated


def cosine_similarity(a, b):
    a = np.array(a, dtype=np.float32)
    b = np.array(b, dtype=np.float32)
    norm_a = np.linalg.norm(a)
    norm_b = np.linalg.norm(b)
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return float(np.dot(a, b) / (norm_a * norm_b))


# ── Face Recognition Core ─────────────────────────────────────────────────────
def recognize_faces_in_image(img: np.ndarray):
    collection = get_collection()
    if collection is None:
        print("[ERROR] ChromaDB collection not found. Run generate_embeddings_improved.py")
        return set(), 0, []

    with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
        tmp_path = tmp.name
    cv2.imwrite(tmp_path, img)

    recognized = set()
    debug_info = []
    face_count = 0

    try:
        faces_data = DeepFace.extract_faces(
            img_path=tmp_path,
            detector_backend=DETECTOR,
            enforce_detection=False,
            align=True,
        )
        face_count = len(faces_data)
        print(f"\n[DEBUG] Faces detected: {face_count}")

        for i, face_obj in enumerate(faces_data):
            face_arr = face_obj.get("face")
            if face_arr is None:
                continue

            face_uint8 = (face_arr * 255).astype(np.uint8)
            face_bgr   = cv2.cvtColor(face_uint8, cv2.COLOR_RGB2BGR)
            with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as ftmp:
                face_tmp = ftmp.name
            cv2.imwrite(face_tmp, face_bgr)

            try:
                emb_objs = DeepFace.represent(
                    img_path=face_tmp,
                    model_name=MODEL_NAME,
                    detector_backend="skip",
                    enforce_detection=False,
                    align=False,
                )
                if not emb_objs:
                    continue

                query_emb = l2_normalize(emb_objs[0]["embedding"])

                results = collection.query(
                    query_embeddings=[query_emb],
                    n_results=1,
                    include=["embeddings", "metadatas", "distances"],
                )

                chroma_distance = results["distances"][0][0]
                meta       = results["metadatas"][0][0]
                student_id = meta["student_id"]

                stored_emb = results["embeddings"][0][0] if results.get("embeddings") else None
                if stored_emb is not None:
                    similarity = cosine_similarity(query_emb, stored_emb)
                else:
                    similarity = round(1.0 - chroma_distance, 4)
                distance = round(1.0 - similarity, 4)

                is_match = similarity >= SIMILARITY_THRESHOLD
                print(f"  Face {i+1}: best_match={student_id}  "
                      f"similarity={similarity:.4f}  distance={distance:.4f}  "
                      f"{'MATCH' if is_match else 'UNKNOWN'}")

                entry = {
                    "face_index": i + 1,
                    "best_match": student_id,
                    "similarity": round(similarity, 4),
                    "distance":   round(distance, 4),
                    "confidence": round(similarity * 100, 1),
                    "matched":    is_match,
                }

                if is_match and student_id not in recognized:
                    recognized.add(student_id)
                    entry["student_id"] = student_id
                else:
                    entry["student_id"] = "Unknown"

                debug_info.append(entry)

            except Exception as e:
                print(f"  Face {i+1}: embedding error — {e}")
            finally:
                if os.path.exists(face_tmp):
                    os.remove(face_tmp)

    except Exception as e:
        print(f"[ERROR] extract_faces failed: {e}")
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    return recognized, face_count, debug_info


RNO_PATTERN = re.compile(r'^\d{2}wh[15]a66\d{2}$', re.IGNORECASE)


def get_device_fingerprint():
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()
    ua = request.headers.get('User-Agent', '')
    return hashlib.sha256(f"{ip}|{ua}".encode()).hexdigest()[:16]


# ── Faculty Auth Endpoints ────────────────────────────────────────────────────
@app.route("/api/auth/register", methods=["POST"])
def register():
    data = request.get_json()
    name, email, password = data.get("name"), data.get("email"), data.get("password")
    if not all([name, email, password]):
        return jsonify({"message": "All fields are required"}), 400
    valid, msg = validate_password(password)
    if not valid:
        return jsonify({"message": msg}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"message": "Database connection failed"}), 500
    cur = conn.cursor()
    cur.execute("SELECT id FROM faculty WHERE email = %s", (email,))
    if cur.fetchone():
        cur.close(); conn.close()
        return jsonify({"message": "Email already registered"}), 400
    cur.execute(
        "INSERT INTO faculty (name, email, password) VALUES (%s, %s, %s)",
        (name, email, generate_password_hash(password)),
    )
    conn.commit(); cur.close(); conn.close()
    return jsonify({"message": "Registration successful"}), 201


@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.get_json()
    email, password = data.get("email"), data.get("password")
    if not all([email, password]):
        return jsonify({"message": "Email and password are required"}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"message": "Database connection failed"}), 500
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM faculty WHERE email = %s", (email,))
    user = cur.fetchone(); cur.close(); conn.close()

    if not user or not check_password_hash(user["password"], password):
        return jsonify({"message": "Invalid email or password"}), 401

    token = jwt.encode(
        {"user_id": user["id"], "email": user["email"],
         "exp": datetime.utcnow() + timedelta(hours=24)},
        app.config["SECRET_KEY"], algorithm="HS256",
    )
    return jsonify({"token": token, "name": user["name"], "email": user["email"]}), 200


# ── Attendance — Image Upload ─────────────────────────────────────────────────
@app.route("/api/attendance/upload", methods=["POST"])
@token_required
def upload_attendance(current_user_id):
    if "images" not in request.files:
        return jsonify({"message": "No images provided"}), 400

    files = request.files.getlist("images")
    if not files:
        return jsonify({"message": "No images selected"}), 400

    if get_collection() is None:
        return jsonify({"message": "Embeddings not found. Run generate_embeddings_improved.py"}), 400

    recognized_students = set()
    total_faces = 0
    all_debug = []

    print(f"\n{'='*60}")
    print(f"Processing {len(files)} image(s)...")

    for idx, file in enumerate(files, 1):
        if not file.filename:
            continue
        file_bytes = np.frombuffer(file.read(), np.uint8)
        img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
        if img is None:
            print(f"[WARN] Image {idx}: failed to decode")
            continue

        recognized, face_count, debug_info = recognize_faces_in_image(img)
        recognized_students.update(recognized)
        total_faces += face_count
        all_debug.extend(debug_info)

        matched = [d["student_id"] for d in debug_info if d["matched"]]
        unknown = sum(1 for d in debug_info if not d["matched"])
        print(f"Image {idx}: {face_count} faces | matched={matched} | unknown={unknown}")

    print(f"\nFINAL: {total_faces} faces detected, {len(recognized_students)} students matched")
    print(f"{'='*60}\n")

    session_id = str(uuid.uuid4())
    conn = get_db_connection()
    if conn:
        cur = conn.cursor()
        ts = datetime.now()
        for sid in recognized_students:
            cur.execute(
                "INSERT INTO attendance (student_id, faculty_id, timestamp, method, session_id) VALUES (%s,%s,%s,'face',%s)",
                (sid, current_user_id, ts, session_id),
            )
        conn.commit(); cur.close(); conn.close()

    unknown_count = max(total_faces - len(recognized_students), 0)
    accuracy = round((len(recognized_students) / total_faces * 100), 1) if total_faces > 0 else 0

    return jsonify({
        "message":       f"Attendance recorded for {len(recognized_students)} students",
        "session_id":    session_id,
        "students":      sorted(recognized_students),
        "total_faces":   total_faces,
        "matched_count": len(recognized_students),
        "unknown_count": unknown_count,
        "accuracy":      accuracy,
        "debug":         all_debug,
    }), 200


# ── QR Attendance ─────────────────────────────────────────────────────────────
@app.route("/api/qr/generate", methods=["POST"])
@token_required
def generate_qr(current_user_id):
    body         = request.get_json(force=True, silent=True) or {}
    frontend_url = (body.get("frontend_url") or "").strip().rstrip("/") or FRONTEND_URL
    backend_url  = (body.get("backend_url")  or "").strip().rstrip("/") or BACKEND_URL

    qr_token   = str(uuid.uuid4())
    session_id = str(uuid.uuid4())
    expires_at = time.time() + QR_TTL_SEC

    QR_STORE[qr_token] = {
        "faculty_id": current_user_id,
        "expires_at": expires_at,
        "session_id": session_id,
        "scanned_by": {},
        "devices":    set(),
    }

    scan_url = f"{frontend_url}/qr/{qr_token}?api={backend_url}"
    qr_img = qrcode.make(scan_url)
    buf = BytesIO()
    qr_img.save(buf, format="PNG")
    buf.seek(0)
    qr_b64 = base64.b64encode(buf.read()).decode()

    print(f"[QR] token={qr_token[:8]}... frontend={frontend_url} backend={backend_url}")

    return jsonify({
        "qr_token":    qr_token,
        "session_id":  session_id,
        "qr_image":    f"data:image/png;base64,{qr_b64}",
        "scan_url":    scan_url,
        "backend_url": backend_url,
        "expires_in":  QR_TTL_SEC,
        "message":     f"QR code valid for {QR_TTL_SEC} seconds",
    }), 200


@app.route("/api/qr/scan", methods=["POST"])
def scan_qr():
    data       = request.get_json(force=True, silent=True) or {}
    qr_token   = (data.get("qr_token") or "").strip()
    student_id = (data.get("student_id") or "").strip().lower()

    if not qr_token or not student_id:
        return jsonify({"message": "qr_token and student_id are required"}), 400

    if not RNO_PATTERN.match(student_id):
        return jsonify({"message": "Invalid RNO format (e.g. 23wh1a6601)"}), 400

    matched_id = next((s for s in ALL_STUDENTS if s.lower() == student_id), None)
    if not matched_id:
        return jsonify({"message": "RNO not found in system"}), 400

    entry = QR_STORE.get(qr_token)
    if not entry:
        return jsonify({"message": "Invalid QR code"}), 400
    if time.time() > entry["expires_at"]:
        QR_STORE.pop(qr_token, None)
        return jsonify({"message": "QR code has expired"}), 400

    if matched_id in entry["scanned_by"]:
        return jsonify({"message": "Attendance already marked for this RNO"}), 400

    fp = get_device_fingerprint()
    if fp in entry["devices"]:
        return jsonify({"message": "This device has already submitted attendance for this session"}), 400

    entry["scanned_by"][matched_id] = datetime.now().strftime("%H:%M:%S")
    entry["devices"].add(fp)
    print(f"[QR] {matched_id} | fp={fp} | total={len(entry['scanned_by'])}")

    conn = get_db_connection()
    if conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO qr_attendance (student_id, qr_token, faculty_id) VALUES (%s,%s,%s)",
            (matched_id, qr_token, entry["faculty_id"]),
        )
        cur.execute(
            "INSERT INTO attendance (student_id, faculty_id, method, session_id) VALUES (%s,%s,'qr',%s)",
            (matched_id, entry["faculty_id"], entry["session_id"]),
        )
        conn.commit(); cur.close(); conn.close()

    return jsonify({
        "message":       f"Attendance marked for {matched_id}",
        "student_id":    matched_id,
        "total_scanned": len(entry["scanned_by"]),
    }), 200


# ── Manual Attendance ─────────────────────────────────────────────────────────
@app.route("/api/attendance/manual", methods=["POST"])
@token_required
def manual_attendance(current_user_id):
    data        = request.get_json(force=True, silent=True) or {}
    student_ids = data.get("student_ids") or []
    if isinstance(student_ids, str):
        student_ids = [student_ids]

    added, invalid = [], []
    conn = get_db_connection()
    if not conn:
        return jsonify({"message": "Database connection failed"}), 500
    cur = conn.cursor()
    ts         = datetime.now()
    session_id = str(uuid.uuid4())

    for raw in student_ids:
        sid = raw.strip().lower()
        matched = next((s for s in ALL_STUDENTS if s.lower() == sid), None)
        if not matched:
            invalid.append(raw)
            continue
        cur.execute(
            "INSERT INTO attendance (student_id, faculty_id, timestamp, method, session_id) VALUES (%s,%s,%s,'manual',%s)",
            (matched, current_user_id, ts, session_id),
        )
        added.append(matched)

    conn.commit(); cur.close(); conn.close()
    print(f"[MANUAL] session={session_id} added={added} invalid={invalid}")
    return jsonify({
        "message":    f"Manually added {len(added)} student(s).",
        "session_id": session_id,
        "added":      added,
        "invalid":    invalid,
    }), 200


# ── QR Status ─────────────────────────────────────────────────────────────────
@app.route("/api/qr/public-status/<qr_token>", methods=["GET"])
def qr_public_status(qr_token):
    entry = QR_STORE.get(qr_token)
    if not entry:
        return jsonify({"active": False, "message": "Invalid or expired QR"}), 404
    remaining = max(0, int(entry["expires_at"] - time.time()))
    return jsonify({"active": remaining > 0, "expires_in": remaining}), 200


@app.route("/api/qr/status/<qr_token>", methods=["GET"])
@token_required
def qr_status(current_user_id, qr_token):
    entry = QR_STORE.get(qr_token)
    if not entry:
        return jsonify({"message": "QR not found or expired"}), 404
    remaining = max(0, int(entry["expires_at"] - time.time()))
    return jsonify({
        "scanned_by":    sorted(entry["scanned_by"].keys()),
        "scanned_list":  [{"student_id": sid, "time": ts} for sid, ts in sorted(entry["scanned_by"].items())],
        "scanned_count": len(entry["scanned_by"]),
        "expires_in":    remaining,
        "active":        remaining > 0,
    }), 200


# ── Report ────────────────────────────────────────────────────────────────────
@app.route("/api/attendance/report", methods=["GET"])
@token_required
def get_report(current_user_id):
    session_ids_raw = request.args.get("session_ids", "").strip()
    method_filter   = request.args.get("method", "all").lower().strip()
    if method_filter not in {"all", "face", "qr", "manual"}:
        method_filter = "all"

    session_ids = [s.strip() for s in session_ids_raw.split(",") if s.strip()]
    if not session_ids:
        return jsonify({"message": "session_ids is required"}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"message": "Database connection failed"}), 500

    placeholders = ",".join(["%s"] * len(session_ids))
    cur = conn.cursor(dictionary=True)
    if method_filter == "all":
        cur.execute(f"""
            SELECT student_id,
                   MIN(timestamp) AS timestamp,
                   GROUP_CONCAT(DISTINCT method ORDER BY method) AS methods
            FROM attendance
            WHERE faculty_id = %s AND session_id IN ({placeholders})
            GROUP BY student_id
            ORDER BY student_id
        """, (current_user_id, *session_ids))
    else:
        cur.execute(f"""
            SELECT student_id,
                   MIN(timestamp) AS timestamp,
                   method AS methods
            FROM attendance
            WHERE faculty_id = %s AND session_id IN ({placeholders}) AND method = %s
            GROUP BY student_id
            ORDER BY student_id
        """, (current_user_id, *session_ids, method_filter))
    records = cur.fetchall()
    cur.close(); conn.close()

    if not records:
        return jsonify({"message": "No attendance records found for this session"}), 404

    present = {
        r["student_id"]: {
            "timestamp": r["timestamp"].strftime("%Y-%m-%d %H:%M:%S") if r["timestamp"] else "-",
            "methods":   r["methods"] or "-",
        }
        for r in records
    }

    rows = []
    for sid in ALL_STUDENTS:
        if sid in present:
            rows.append({"Student_ID": sid, "Status": "Present",
                         "Method": present[sid]["methods"], "Timestamp": present[sid]["timestamp"]})
        else:
            rows.append({"Student_ID": sid, "Status": "Absent", "Method": "-", "Timestamp": "-"})

    df = pd.DataFrame(rows)
    excel_path = os.path.join(DATA_DIR, "attendance_report.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Attendance")
        ws = writer.sheets["Attendance"]
        for col, width in [("A", 15), ("B", 12), ("C", 20), ("D", 22)]:
            ws.column_dimensions[col].width = width
        from openpyxl.styles import PatternFill
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            fill = green if row[1].value == "Present" else red
            for cell in row:
                cell.fill = fill

    return send_file(excel_path, as_attachment=True,
                     download_name=f"attendance_report_{method_filter}.xlsx")


@app.route("/api/attendance/stats", methods=["GET"])
@token_required
def get_attendance_stats(current_user_id):
    session_ids_raw = request.args.get("session_ids", "").strip()
    session_ids = [s.strip() for s in session_ids_raw.split(",") if s.strip()]
    total = len(ALL_STUDENTS)
    present = 0
    if session_ids:
        conn = get_db_connection()
        if conn:
            placeholders = ",".join(["%s"] * len(session_ids))
            cur = conn.cursor()
            cur.execute(f"""
                SELECT COUNT(DISTINCT student_id) FROM attendance
                WHERE faculty_id = %s AND session_id IN ({placeholders})
            """, (current_user_id, *session_ids))
            present = cur.fetchone()[0] or 0
            cur.close(); conn.close()
    return jsonify({"total": total, "present": present, "absent": total - present}), 200


@app.route("/api/students/list", methods=["GET"])
@token_required
def get_students_list(current_user_id):
    return jsonify({"students": ALL_STUDENTS, "total": len(ALL_STUDENTS)}), 200


@app.route("/api/embeddings/validate", methods=["GET"])
@token_required
def validate_embeddings(current_user_id):
    col = get_collection()
    if col is None:
        return jsonify({"valid": False, "message": "Run generate_embeddings_improved.py"}), 400
    all_data = col.get()
    stored = {m["student_id"] for m in all_data.get("metadatas", [])}
    missing = sorted(set(ALL_STUDENTS) - stored)
    return jsonify({
        "valid":             len(missing) == 0,
        "total_students":    len(ALL_STUDENTS),
        "stored_students":   len(stored),
        "total_embeddings":  len(all_data.get("ids", [])),
        "missing_students":  missing,
    }), 200


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5000, debug=True)
